"""
Construye ../webapp/public/data/precios_fantastica.json a partir del Excel de
precios de consenso de la liga (../webapp/public/data/*.xlsx).

Por qué un script separado de main.py/build_catalog.py: esos dos regeneran
players.json y players/<id>.json entero cada vez que se ejecutan (pisando
cualquier cosa que se les añada a mano), así que el precio Fantástica -que
es un dato manual, pactado por la liga y no algo que dé la API de Marca-
tiene que vivir en su propio fichero para sobrevivir a esas regeneraciones.

El Excel no trae el id de fantasy.marca.com de cada jugador (por algo lo
gestionan aparte de la propia liga), así que hay que casarlo por nombre
contra el catálogo. El fichero tiene un formato "ancho": una columna de
posición (Portero/Defensa/Medio/Delantero) que aplica a toda la fila, y
luego un bloque de 3 columnas (nombre, precio, separador) por equipo. Los
nombres del Excel suelen ser abreviados ("J. Owono", "Fdez.", "Glez.") así
que el emparejamiento es aproximado (fuzzy) y se restringe a jugadores del
mismo equipo+posición para no confundir gente.

Ese emparejamiento aproximado falla en un puñado de casos reales -jugadores
nuevos que el Excel ya incluye pero que fantasy.marca.com todavía no ha
dado de alta, o entradas que sencillamente no tienen contrapartida clara-.
MANUAL_OVERRIDES resuelve esos casos a mano (revisado contra el catálogo
completo, ver conversación de origen). Si en una actualización futura del
Excel aparecen jugadores nuevos sin match, el script los lista al final:
hay que decidir si son fichajes que Marca aún no ha añadido (se ignoran,
sin más, hasta la siguiente actualización) o si hace falta añadir un
override.

Uso:
    python build_precios_fantastica.py
"""

import difflib
import glob
import json
import os
import re
import sys
import unicodedata

import openpyxl

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "webapp", "public", "data")

# Alias del nombre de equipo tal cual aparece en el Excel -> id_team de teams.json.
# Si el Excel de una temporada futura trae equipos distintos (ascensos/descensos),
# esta tabla hay que actualizarla a mano.
TEAM_ALIAS = {
    "Alavés": 48,
    "Athletic": 1,
    "At. Madrid": 2,
    "Barcelona": 3,
    "Betis": 4,
    "Celta": 5,
    "Dep. Coruña": 6,
    "Elche": 23,
    "Español": 8,
    "Getafe": 9,
    "Levante": 12,
    "Málaga": 13,
    "Osasuna": 50,
    "Racing Santander": 1490,
    "Rayo Vallecano": 14,
    "R. Madrid": 15,
    "R. Sociedad": 16,
    "Sevilla": 17,
    "Valencia": 19,
    "Villarreal": 20,
}

POSITION_BY_LABEL = {"Portero": 1, "Defensa": 2, "Medio": 3, "Delantero": 4}

# Abreviaturas de apellido habituales en el Excel que el catálogo escribe completas.
SURNAME_ABBREVIATIONS = {
    "fdez": "fernandez",
    "glez": "gonzalez",
    "hdez": "hernandez",
    "mtnez": "martinez",
    "rguez": "rodriguez",
}

# (equipo excel, posición excel, nombre excel) -> id de fantasy.marca.com, o None
# para marcar explícitamente "sin contrapartida en el catálogo, no forzar match".
# Revisado a mano para el Excel "Listado Jugadores 26-27" (ver histórico del repo).
MANUAL_OVERRIDES = {
    ("Racing Santander", "Portero", "Laro Gómez"): None,
    ("Villarreal", "Portero", "Péter Gulácsi"): None,
    ("Celta", "Defensa", "Abdoulaye Faye"): None,
    ("Racing Santander", "Defensa", "P. Felipe"): None,
    ("Racing Santander", "Defensa", "P. Ramón"): None,
    ("Español", "Defensa", "Roger Hinojo"): None,
    ("Español", "Defensa", "Unai Núñez"): None,
    ("Málaga", "Defensa", "J. Salinas"): None,
    ("Getafe", "Defensa", "Sazonov"): None,
    ("Elche", "Medio", "J. Morcillo"): None,
    ("Elche", "Delantero", "Fer Niño"): None,
    ("Elche", "Delantero", "U. Konare"): None,
    ("Osasuna", "Delantero", "Dubasin"): None,
    ("Rayo Vallecano", "Defensa", "Kumbulla"): None,
    ("Sevilla", "Medio", "Miguel Sierra"): None,
    ("Sevilla", "Medio", "P. Mercado"): None,
    ("Levante", "Portero", "Mathew Ryan"): None,
    ("Barcelona", "Medio", "Jesse Bisiwu"): None,
    ("Athletic", "Medio", "Generabarrena"): None,
    ("Celta", "Portero", "A. Bayindir"): None,
    ("Racing Santander", "Delantero", "Yassir Zabiri"): None,
    ("Celta", "Medio", "Hugo Glez."): None,
    ("Dep. Coruña", "Defensa", "Angeliño"): None,
    ("Sevilla", "Portero", "Fran Glez."): None,
    ("Levante", "Delantero", "Yanis Musuayi"): None,
    ("Sevilla", "Defensa", "Julio Díaz"): None,
    # Error de columna en el propio Excel: esta fila cae bajo el bloque de
    # "R. Madrid" pero el nombre solo existe en el catálogo como delantero
    # del Levante (id 63780), que si no se queda sin precio.
    ("R. Madrid", "Delantero", "Carlos Espí"): 63780,
}


def find_excel_path() -> str:
    matches = glob.glob(os.path.join(DATA_DIR, "*.xlsx"))
    if len(matches) != 1:
        raise SystemExit(
            f"Esperaba exactamente un .xlsx en {DATA_DIR}, encontrados {len(matches)}: {matches}"
        )
    return matches[0]


def normalize_tokens(name: str) -> list[str]:
    ascii_name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    ascii_name = ascii_name.lower().replace("-", " ")
    ascii_name = re.sub(r"[^a-z0-9. ]", "", ascii_name)
    tokens = []
    for tok in ascii_name.split():
        tok = tok.rstrip(".")
        tokens.append(SURNAME_ABBREVIATIONS.get(tok, tok))
    return tokens


def token_score(e_tokens: list[str], c_tokens: list[str]) -> float:
    total_weight = total_score = 0.0
    for tok in e_tokens:
        weight = 1.0 if len(tok) > 1 else 0.4  # una inicial sola pesa menos que un nombre completo
        if len(tok) == 1:
            best = 1.0 if any(c.startswith(tok) for c in c_tokens) else 0.0
        else:
            best = max((difflib.SequenceMatcher(None, tok, c).ratio() for c in c_tokens), default=0.0)
        total_score += weight * best
        total_weight += weight
    return total_score / total_weight if total_weight else 0.0


def parse_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Precios Jugadores"] if "Precios Jugadores" in wb.sheetnames else wb.worksheets[0]

    team_cols = {}
    for col in range(2, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            team_cols[col] = value

    unknown_teams = sorted(set(team_cols.values()) - set(TEAM_ALIAS))
    if unknown_teams:
        raise SystemExit(
            f"Equipos del Excel sin alias en TEAM_ALIAS: {unknown_teams}. "
            "Añádelos a TEAM_ALIAS (o revisa si son un renombrado de uno existente)."
        )

    records = []
    current_position = None
    for row in range(2, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if label:
            current_position = label
        for col, team in team_cols.items():
            name = ws.cell(row=row, column=col).value
            price = ws.cell(row=row, column=col + 1).value
            if name:
                records.append({"team": team, "position": current_position, "name": name, "price": price})
    return records


def match_records(records: list[dict], players: list[dict]) -> tuple[dict, list[dict]]:
    players_by_group = {}
    for p in players:
        players_by_group.setdefault((p["id_team"], p["position"]), []).append(p)

    price_by_id: dict[int, float] = {}
    unmatched: list[dict] = []
    reserved_player_ids: set[int] = set()  # ids ya resueltos a mano, no elegibles para el matching automático

    # Los overrides manuales se aplican primero, y reservan su id de destino
    # para que el matching automático no se lo pueda quedar por error en otra fila.
    auto_indices = []
    for idx, rec in enumerate(records):
        override_key = (rec["team"], rec["position"], rec["name"])
        if override_key in MANUAL_OVERRIDES:
            override_id = MANUAL_OVERRIDES[override_key]
            if override_id is None:
                unmatched.append(rec)
            else:
                price_by_id[override_id] = rec["price"]
                reserved_player_ids.add(override_id)
        elif rec["price"] is None:
            # Fila con nombre pero sin precio puesto todavía en el Excel: no hay
            # nada que asignar, y no debe competir por un candidato con otra fila
            # que sí tenga precio.
            unmatched.append(rec)
        else:
            auto_indices.append(idx)

    groups = {}
    for idx in auto_indices:
        rec = records[idx]
        key = (TEAM_ALIAS[rec["team"]], POSITION_BY_LABEL[rec["position"]])
        groups.setdefault(key, []).append(idx)

    for key, rec_indices in groups.items():
        candidates = [c for c in players_by_group.get(key, []) if c["id"] not in reserved_player_ids]
        pairs = []
        for idx in rec_indices:
            e_tokens = normalize_tokens(records[idx]["name"])
            for c in candidates:
                pairs.append((token_score(e_tokens, normalize_tokens(c["name"])), idx, c["id"]))
        pairs.sort(key=lambda item: -item[0])

        claimed_records, claimed_players = set(), set()
        for score, idx, player_id in pairs:
            if idx in claimed_records or player_id in claimed_players:
                continue
            claimed_records.add(idx)
            claimed_players.add(player_id)
            price_by_id[player_id] = records[idx]["price"]

        for idx in rec_indices:
            if idx not in claimed_records:
                unmatched.append(records[idx])

    return price_by_id, unmatched


def main() -> None:
    excel_path = find_excel_path()
    print(f"Leyendo {excel_path}...", file=sys.stderr)
    records = parse_excel(excel_path)
    print(f"  {len(records)} filas de jugador", file=sys.stderr)

    players_path = os.path.join(DATA_DIR, "players.json")
    with open(players_path, encoding="utf-8") as f:
        players = json.load(f)

    price_by_id, unmatched = match_records(records, players)

    output = {str(pid): round(price, 2) for pid, price in sorted(price_by_id.items())}

    out_path = os.path.join(DATA_DIR, "precios_fantastica.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
        f.write("\n")

    print(f"\nGuardado {out_path}: {len(output)} jugadores con precio Fantástica.", file=sys.stderr)

    if unmatched:
        print(f"\n{len(unmatched)} filas del Excel sin match en el catálogo (sin precio asignado):", file=sys.stderr)
        for rec in sorted(unmatched, key=lambda r: (r["team"], r["position"])):
            print(f"  {rec['team']:18s} {rec['position']:10s} {rec['name']}", file=sys.stderr)
        print(
            "  -> normalmente son fichajes de este verano que fantasy.marca.com "
            "todavía no ha dado de alta; si alguno lleva tiempo sin aparecer, "
            "revisa MANUAL_OVERRIDES.",
            file=sys.stderr,
        )


if __name__ == "__main__":
    main()
